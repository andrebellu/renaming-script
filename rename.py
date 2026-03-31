import os
from pprint import pprint
import pandas as pd
from PIL import Image
from PIL.ExifTags import TAGS
import datetime
import shutil
import time
import argparse
import sys
import openpyxl
from pillow_heif import register_heif_opener

register_heif_opener()

# Device configurations
DEVICES = {
    'samsung': ('./photos/samsung', 'SamsungS23'),
    'iphone': ('./photos/iphone', 'iPhone')
}

# Default configurations
DEFAULT_EXCEL_PATH = 'rename_test.xlsx'
DEFAULT_OUTPUT_EXCEL_PATH = 'new_db.xlsx'
DEFAULT_SHEET_NAME = 'Data'


def exif_extraction(file_path) -> list:
  exif_dict = []
  exif_dict.append(os.path.basename(file_path))
  # print(exif_dict)
  try:
    pil_img = Image.open(file_path)
    exif_raw = pil_img.getexif()
    exif_info = {}
    if exif_raw is not None:
        exif_info.update(exif_raw)
        if hasattr(exif_raw, 'get_ifd'):
            exif_info.update(exif_raw.get_ifd(0x8769))
    if exif_info:
      exif = {TAGS.get(k, k): v for k, v in exif_info.items()}

      if "DateTimeOriginal" in exif:
        dt_obj = datetime.datetime.strptime(exif["DateTimeOriginal"], '%Y:%m:%d %H:%M:%S')
        exif_dict.append(dt_obj.timestamp())

      elif "DateTime" in exif:
        dt_obj = datetime.datetime.strptime(exif["DateTime"], '%Y:%m:%d %H:%M:%S')
        exif_dict.append(dt_obj.timestamp())

      if "FocalLength" in exif:
        exif_dict.append(exif["FocalLength"])

  except Exception as e:
    print(f"Warning: Could not read EXIF data for {file_path}. Error: {e}")
    pass

  return exif_dict

def bulk_rename(path_options, excel_path, sheet_name='Data') -> list:
    exif_details_map = {}
    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    valid_rows = df.loc[df['Modello_Telefono'] == path_options[1]]

    target_names = valid_rows['ID_Foto'].dropna().tolist()
    print(f"ID_Foto from Excel: {target_names}")

    valid_extensions = ('.jpg', '.jpeg', '.png', '.JPG', '.HEIC', '.heic')
    all_files_in_dir = [f for f in os.listdir(path_options[0]) if f.endswith(valid_extensions) and not f.startswith('.')]

    files_with_exif = []
    files_without_exif = []

    print("Processing files to check EXIF data...")
    for filename in all_files_in_dir:
        file_path = os.path.join(path_options[0], filename)
        exif_result = exif_extraction(file_path)
        original_filename = exif_result[0]

        acquisition_time = None
        focal_length = None

        if len(exif_result) > 1:
            acquisition_time = exif_result[1]
        if len(exif_result) > 2:
            focal_length = exif_result[2]

        exif_details_map[original_filename] = {'timestamp': acquisition_time, 'focal_length': focal_length}

        if acquisition_time is not None:
            files_with_exif.append((filename, acquisition_time))
        else:
            files_without_exif.append(filename)

    # Sort files that have EXIF data by their acquisition time
    files_with_exif.sort(key=lambda x: x[1])
    sorted_filenames = [f[0] for f in files_with_exif]

    if files_without_exif:
        print(f"The following {len(files_without_exif)} files could not be processed due to missing or problematic EXIF data and require manual handling:")
        for f in files_without_exif:
            print(f" - {f}")

    if len(sorted_filenames) != len(target_names):
        print(f"ERROR: ({len(sorted_filenames)}) files with valid EXIF != ({len(target_names)}) excel ids")
        return []

    confirm = input("Proceed with renaming files with valid EXIF data? (y/n) ")
    if confirm.lower() != 'y':
        return []

    os.makedirs(os.path.join(path_options[0], "renamed"), exist_ok=True)

    excel_update_data = []

    for i, filename in enumerate(sorted_filenames):
        original_filename = filename
        extension = os.path.splitext(filename)[1].lower()
        new_name_base = target_names[i]
        new_filename = os.path.splitext(new_name_base)[0] + extension

        exif_data_for_original = exif_details_map.get(original_filename, {})
        focal_length_for_excel = exif_data_for_original.get('focal_length')

        excel_update_data.append({
            'ID_Foto': new_filename,
            'Focale_EXIF': str(focal_length_for_excel) if focal_length_for_excel is not None else None
        })

        old_path = os.path.join(path_options[0], filename)
        new_path = os.path.join(path_options[0], "renamed", new_filename)

        try:
            shutil.copy2(old_path, new_path)
            # print(f"Copiato e rinominato: {filename} -> {new_filename}")
        except Exception as e:
            print(f"Error on {filename}: {e}")
    print(excel_update_data)
    return excel_update_data

def update_excel_with_exif_data(excel_path, output_excel_path, sheet_name, exif_data) -> None:
    print(exif_data)
    
    if not output_excel_path or output_excel_path == excel_path:
        backup_path = f"{excel_path}.bak"
        try:
            shutil.copy2(excel_path, backup_path)
            print(f"Backup: {backup_path}")
        except Exception as e:
            print(f"Error during backup creation: {e}")
            return
        target_path = excel_path
    else:
        try:
            if not os.path.exists(output_excel_path):
                shutil.copy2(excel_path, output_excel_path)
            target_path = output_excel_path
        except Exception as e:
            print(f"Errore durante la gestione del file di output: {e}")
            return

    # openpyxl instead of xlsxwriter
    try:
        wb_values = openpyxl.load_workbook(target_path, data_only=True)
        wb = openpyxl.load_workbook(target_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"Error: The sheet '{sheet_name}' does not exist in the file.")
            return
            
        ws_values = wb_values[sheet_name]
        ws = wb[sheet_name]
        
        header_row = 1
        id_foto_col = None
        focale_col = None
        
        for cell in ws_values[header_row]:
            cell_value = str(cell.value).strip() if cell.value else ""
            if cell_value == 'ID_Foto':
                id_foto_col = cell.column
            elif cell_value == 'Focale_EXIF':
                focale_col = cell.column
                
        if not id_foto_col or not focale_col:
            print("Error: Unable to find 'ID_Foto' or 'Focale_EXIF' columns in the header.")
            return

        for entry in exif_data:
            new_filename = str(entry.get('ID_Foto')).strip()
            focal_length = entry.get('Focale_EXIF')
            
            if new_filename and focal_length is not None:
                found = False
                for row in range(2, ws_values.max_row + 1):
                    cell_val = str(ws_values.cell(row=row, column=id_foto_col).value).strip()
                    if os.path.splitext(cell_val)[0] == os.path.splitext(new_filename)[0]:
                        ws.cell(row=row, column=focale_col).value = focal_length
                        found = True
                        break
                if not found:
                    print(f"Warning: ID_Foto '{new_filename}' not found. No update performed.")
        
        wb.save(target_path)
        print(f"Excel file '{target_path}' updated successfully.")
        
    except Exception as e:
        print(f"Error during update with openpyxl: {e}")

def exif_info(file_path):
    try:
        pil_img = Image.open(file_path)
        exif_raw = pil_img.getexif()
        exif_info = {}
        if exif_raw is not None:
            exif_info.update(exif_raw)
            if hasattr(exif_raw, 'get_ifd'):
                exif_info.update(exif_raw.get_ifd(0x8769))
        if exif_info:
            exif = {TAGS.get(k, k): v for k, v in exif_info.items()}
            return exif
    except Exception as e:
        print(f"Warning: Could not read EXIF data for {file_path}. Error: {e}")
    return None

def main():
    parser = argparse.ArgumentParser(
        description='Photo renaming and EXIF data extraction tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python rename.py --action bulk-rename --device iphone
  python rename.py --action bulk-rename --device samsung --output output.xlsx
  python rename.py --action exif-info --file photos/image.jpg
        '''
    )
    
    parser.add_argument(
        '--action',
        choices=['bulk-rename', 'exif-info'],
        required=True,
        help='Action to perform'
    )
    
    parser.add_argument(
        '--device',
        choices=['samsung', 'iphone'],
        help='Device to process (required for bulk-rename)'
    )
    
    parser.add_argument(
        '--file',
        help='File path for exif-info action'
    )
    
    parser.add_argument(
        '--excel',
        default=DEFAULT_EXCEL_PATH,
        help=f'Excel file path (default: {DEFAULT_EXCEL_PATH})'
    )
    
    parser.add_argument(
        '--output',
        default=DEFAULT_OUTPUT_EXCEL_PATH,
        help=f'Output Excel file path (default: {DEFAULT_OUTPUT_EXCEL_PATH})'
    )
    
    parser.add_argument(
        '--sheet',
        default=DEFAULT_SHEET_NAME,
        help=f'Excel sheet name (default: {DEFAULT_SHEET_NAME})'
    )
    
    args = parser.parse_args()
    
    start_time = time.time()
    
    try:
        if args.action == 'bulk-rename':
            if not args.device:
                parser.error('--device is required for bulk-rename action')
            
            if args.device not in DEVICES:
                parser.error(f'Unknown device: {args.device}')
            
            path_options = DEVICES[args.device]
            print(f"Processing {args.device.upper()} photos...")
            
            rows = bulk_rename(path_options, args.excel, args.sheet)
            if rows:
                print(f"Updating Excel file: {args.output}")
                update_excel_with_exif_data(args.excel, args.output, args.sheet, rows)
            else:
                print("No data to update Excel file.")
        
        elif args.action == 'exif-info':
            if not args.file:
                parser.error('--file is required for exif-info action')
            
            print(f"Extracting EXIF data from: {args.file}")
            exif_data = exif_info(args.file)
            if exif_data:
                pprint(exif_data)
            else:
                print("No EXIF data found or file could not be read.")
    
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    
    finally:
        elapsed_time = time.time() - start_time
        print(f"\n--- Completed in {elapsed_time:.2f} seconds ---")

if __name__ == '__main__':
    main()